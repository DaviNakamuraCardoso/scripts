import openpyxl
import pprint

wb = openpyxl.load_workbook('example_sheet_00.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet['A1'])
print(sheet['A1'].value)
b1 = sheet.cell(row=1, column=2)
rows = 6
columns = 3
print(b1)
