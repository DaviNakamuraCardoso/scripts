import openpyxl
import pprint
print('Abrindo workbook...')

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
county_data = {}
rows = sheet.max_row
columns = sheet.max_column
print('Lendo linhas...')
for row in range(2, rows+1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    county_data.setdefault(state, {})
    county_data[state].setdefault(county, {'tracts':0, 'pop': 0})
    county_data[state][county]['tracts'] += 1
    county_data[state][county]['pop'] += int(pop)

print('Escrevendo resultados...')
result_file = open('census_2010.py', 'w')
result_file.write('all_data = ' + pprint.pformat(county_data))
result_file.close()
print('Pronto.')
