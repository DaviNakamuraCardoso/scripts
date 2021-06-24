import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)
wb.remove(wb['Middle Sheet'])
print(wb.sheetnames)
# Or:
del wb['Sheet1']
print(wb.sheetnames)
sheet = wb['First Sheet']
sheet['A1'] = 'É DAVI'
print(sheet['A1'].value)
for i in range(1, 100):
    for j in range(1, 20):
        sheet[openpyxl.utils.get_column_letter(i) + str(j)] = i + j
wb.save('edit_text_wb.xlxs')


