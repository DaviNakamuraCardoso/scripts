import os
import openpyxl
import csv


def convert_to_csv(directory):
    """
    :param directory: The directory of files you want to convert
    :return: The files in csv format
    """
    os.chdir(directory)
    for filename in os.listdir():
        if filename.endswith('.xlsx'):
            filename_file = openpyxl.load_workbook(filename)
            sheets = filename_file.sheetnames
            for sheet in sheets:
                sheet = filename_file[sheet]
                csv_filename = filename.strip('.xlsx') + str(sheet.title) + '.csv'
                csv_file = open(csv_filename, 'w', newline='')
                csv_writer = csv.writer(csv_file)
                for i in range(1, sheet.max_row+1):
                    data = []
                    for j in range(1, sheet.max_column+1):
                        data.append(sheet[openpyxl.utils.get_column_letter(j) + str(i)].value)
                csv_writer.writerow(data)
            csv_file.close()


convert_to_csv('../C13-Planilhas-Excel')