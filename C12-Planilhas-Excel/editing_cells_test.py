import openpyxl
from openpyxl.styles import Font

wb_formulas = openpyxl.load_workbook('redefined_file.xlsx')
sheet = wb_formulas.active
print(sheet.max_row)
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet['A1'] = 'Tall row'
new_sheet['B2'] = 'Wide column'
new_sheet.row_dimensions[1].height = 70
new_sheet.column_dimensions['B'].width = 20
new_sheet.merge_cells('A3:D5')
new_sheet['A3'] = 'Twelve cells merged together'
new_sheet.merge_cells('C6:D6')
new_sheet['C6'] = 'Two cells merged together'
new_wb.save('dimensions.xlsx')

