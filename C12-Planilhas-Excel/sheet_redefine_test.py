#! /usr/bin/python3

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
print('Abrindo arquivo...')
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

price_updates = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}
rows = sheet.max_row
columns = sheet.max_column
bold_font = Font(bold=True)

for i in range(2, rows):
    produce_name = sheet.cell(row=i, column=1).value
    print('Lendo linha', str(i) + '...')
    for j in range(1, 5):
        print('Lendo coluna', str(j) + '...')
        if produce_name in price_updates.keys():
            new_sheet[openpyxl.utils.get_column_letter(j) + str(i)].font = bold_font
            new_sheet['A'+str(i)].value = sheet['A' + str(i)].value
            new_sheet['B' + str(i)].value = price_updates[produce_name]
            new_sheet['C' + str(i)].value = sheet['C' + str(i)].value
            new_sheet['D' + str(i)].value = price_updates[produce_name] * (sheet['C' + str(i)].value)
        else:
            new_sheet[openpyxl.utils.get_column_letter(j) + str(i)].value = sheet[openpyxl.utils.get_column_letter(j) + str(i)].value
last_row = str(rows)
new_sheet['E' + str(rows+1)] = '=SUM(D2:D%s)' % last_row
total_font = Font(name='Times New Roman', size=15, bold=True, italic=True)
new_sheet['E' + str(rows+1)].font = total_font
print('Salvando novo arquivo...')
new_wb.save('redefined_file.xlxs')
print('Pronto!')