import openpyxl
import pprint

wb = openpyxl.load_workbook('example_sheet_00.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet['A1'])
print(sheet['A1'].value)
b1 = sheet.cell(row=1, column=2)
rows = sheet.max_row
columns = sheet.max_column
print(b1)
for i in range(rows):
    for j in range(columns):
        print(sheet.cell(row=i+1, column=j+1).value)

print(openpyxl.utils.column_index_from_string('A'))
print(openpyxl.utils.get_column_letter(48))
matrix = tuple(sheet['A1': (sheet.cell(row=rows, column=columns).coordinate)])
for row_cell_objects in matrix:
    for cell_object in row_cell_objects:
        print(cell_object.coordinate, cell_object.value)
    print('---END OF ROW---')

print(sheet.columns[1])