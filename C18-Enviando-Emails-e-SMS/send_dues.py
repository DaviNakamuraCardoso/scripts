import openpyxl
import smtplib

# smtp_object = smtplib.SMTP('smtp.example.com', 587)
# smtp_object.ehlo()
# smtp_object.starttls()
# smtp_object.login('my_email@example.com', 'my_password')
filename = openpyxl.load_workbook('duesRecords.xlsx')
sheet = filename.active
for i in range(2, sheet.max_row+1):
    months_not_paid = []
    name = sheet['A' + str(i)].value
    email = sheet['B' + str(i)].value
    for j in range(3, sheet.max_column+1):
        if sheet.cell(row=i, column=j).value != 'paid':
            months_not_paid.append(sheet.cell(row=1, column=j).value)
        else:
            continue
    if months_not_paid != []:
        print('Sending', name, 'a email')
        print('From: my_email@example.com')
        print('To:', email)
        print('Subject: ' + name + ", it's time to pay!")
        print('Dear ' + name + ",\nYou didn't paid the following months:")
        # smtp_object.sendmail('my_email@example.com', email, body)
        for month in months_not_paid:
            print(month)

#smtp_object.quit()


