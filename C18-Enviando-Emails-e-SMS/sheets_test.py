import openpyxl
import ezgmail


ss = openpyxl.load_workbook('paiments.xlsx')
sheet = ss.active
for i in range(2, sheet.max_row+1): 
    months_not_paid = []
    name = sheet['A' + str(i)].value
    email = sheet['B' + str(i)].value
    for j in range(3, sheet.max_column+1):
        status = sheet.cell(row=i, column=j).value
        if status != 'Ok':
            months_not_paid.append(sheet.cell(row=1, column=j).value)
    if len(months_not_paid) > 0:
        months = '\n'.join(months_not_paid)
        print("Sending to", name, "email: ", email, '\n', months)
        message = 'Senhor(a) ' + name + ',\n' + ' o senhor(a) se esqueceu de pagar sua mensalidade nos seguinters meses: \n' + months
        ezgmail.send(email, 'TESTE', message, ['/home/davi/Downloads/melado.jpg'])


